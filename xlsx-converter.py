import openpyxl
import pandas as pd

path = input("Enter the full path of the xlsx file: ")
wb = openpyxl.load_workbook(path) #input voor gebruiker

sheet = wb.worksheets[0]
row = sheet[1]

# Rename
sheet.title = 'Details'

#Change Row 1 to BOLD and Color
for cell in row:
    cell.font = openpyxl.styles.Font(bold=True)
    cell.fill = openpyxl.styles.PatternFill("solid", fgColor="43A3B1")

#Hide column B
sheet.column_dimensions['B'].hidden=True

#Grab data from sheet
data=[]
for row in sheet.iter_rows(values_only=True):
    data.append(row)

df = pd.DataFrame(data[1:], columns=data[0])

# Create the pivot table
pivot_table = pd.pivot_table(df, values=['count','price_incl'], index='item_description', aggfunc=sum)

# Sort pivot table on price_incl high to low
pivot_table = pivot_table.sort_values(by='price_incl', ascending=False)


# Create a new sheet in the workbook and save the pivot table data to it
new_sheet = wb.create_sheet(title='Resumé')
for row in pivot_table.iterrows():
    new_sheet.append([row[0]] + list(row[1]))

# Set last row
last_row = new_sheet.max_row

# Add total at bottom
new_sheet[f'A{last_row+1}'] = "Total"

# Sum column B and add last row
col_sum = sum(new_sheet[f'B{i}'].value if new_sheet[f'B{i}'].value is not None else 0 for i in range(1, last_row+1))
new_sheet[f'B{last_row + 1}'] = col_sum

# Sum column C and add last row
col_sum = sum(new_sheet[f'C{i}'].value if new_sheet[f'C{i}'].value is not None else 0 for i in range(1, last_row+1))
new_sheet[f'C{last_row + 1}'] = f"€ {round(col_sum)}"

# Add Euro for whole column C
for i in range(1, last_row + 1):
    if new_sheet[f'C{i}'].value is not None:
        new_sheet[f'C{i}'].value = f"€ {round(new_sheet[f'C{i}'].value, 2)}"

# Create new row in new_sheet
new_sheet.insert_rows(1)
row = new_sheet[1]

#Change Row 1 to BOLD and Color
for cell in row:
    cell.font = openpyxl.styles.Font(bold=True)
    cell.fill = openpyxl.styles.PatternFill("solid", fgColor="43A3B1")

new_sheet['A1'] = "Product"
new_sheet['B1'] = "Sum of Count"
new_sheet['C1'] = "Sum of Price"

# Get a list of all the sheets in the workbook
sheets = wb.sheetnames

# Move a sheet to a new position
wb.move_sheet(wb[sheets[0]], 1)

# Grab data from cell A2
cell = sheet.cell(row=2,column=1)
cell_value = cell.value

#Save to new file
wb.save(cell_value+'_manifest.xlsx')
